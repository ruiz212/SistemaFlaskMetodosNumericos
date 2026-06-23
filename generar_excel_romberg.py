import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import numpy as np
import math

# Sample realistic data from the app (using a bell curve for irradiance)
def rad(hora):
    # Simulated solar radiation with peak at 12:00
    if 6 <= hora <= 18:
        return 800 * math.exp(-((hora - 12) ** 2) / (2 * 2**2))
    return 0.0

# 24 hour points
x_vals = np.arange(24)
y_vals = np.array([rad(x) for x in x_vals])

def f_real(t):
    return float(np.interp(t, x_vals, y_vals))

# Setup Excel Workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Metodo de Romberg"

# Header styling
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
align_center = Alignment(horizontal="center", vertical="center")

headers = [
    "i", "t (Hora)", "f(t) (W/m²)", 
    "j (Nivel)", "Intervalos (n)", "h (Paso)", "k=1 (Trapecio)", 
    "k=2", "k=3", "k=4", "k=5 (Alta Precisión)"
]

ws.append(headers)

for col in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = align_center

# 1. GENERATE DATA POINTS for 16 intervals (5 levels)
# n = 16 intervals, points = 17. Total time = 23.
n = 16
a = 0
b = 23
h_min = (b - a) / n

# Points: i = 0 to 16
for i in range(17):
    t_val = a + i * h_min
    ft = f_real(t_val)
    ws.append([i, t_val, ft])

# Format data
for row in range(2, 19):
    ws.cell(row=row, column=2).number_format = '0.0000'
    ws.cell(row=row, column=3).number_format = '0.0000'

# 2. ROMBERG TABLE (start at row 2, column 4)
# Levels: 1 to 5
for j in range(1, 6):
    row = j + 1
    n_j = 2 ** (j - 1)
    h_j = (b - a) / n_j
    ws.cell(row=row, column=4, value=j)      # j (Nivel)
    ws.cell(row=row, column=5, value=n_j)    # Intervalos (n)
    ws.cell(row=row, column=6, value=h_j)    # h (Paso)
    
    # Formula for Trapecio (k=1)
    # The sum uses data points from column C (rows 2 to 18)
    if j == 1:
        f_trap = f"=(F{row}/2)*(C2 + C18)"
    elif j == 2:
        f_trap = f"=(F{row}/2)*(C2 + 2*C10 + C18)"
    elif j == 3:
        f_trap = f"=(F{row}/2)*(C2 + 2*(C6+C10+C14) + C18)"
    elif j == 4:
        f_trap = f"=(F{row}/2)*(C2 + 2*(C4+C6+C8+C10+C12+C14+C16) + C18)"
    elif j == 5:
        f_trap = f"=(F{row}/2)*(C2 + 2*SUM(C3:C17) + C18)"
        
    ws.cell(row=row, column=7, value=f_trap)

# 3. RICHARDSON EXTRAPOLATION (k=2 to 5)
cols = ['G', 'H', 'I', 'J', 'K']

for k in range(2, 6):
    col_idx = 6 + k  # k=2 -> 8 (H)
    factor = 4 ** (k - 1)
    
    for j in range(1, 6 - k + 1):
        row = j + 1
        prev_col = cols[k - 2]
        f_extrap = f"=({factor}*{prev_col}{row + 1} - {prev_col}{row})/{factor - 1}"
        ws.cell(row=row, column=col_idx, value=f_extrap)

# Apply number formatting to table
for row in range(2, 7):
    ws.cell(row=row, column=6).number_format = '0.0000'
    for col in range(7, 12):
        ws.cell(row=row, column=col).number_format = '0.0000'

# Adjust column widths
for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
    ws.column_dimensions[col].width = 16

# Add explanation
ws.cell(row=20, column=2, value="Explicación (Método Numérico de Romberg):").font = Font(bold=True)
ws.cell(row=21, column=2, value="1. La columna 'k=1' calcula la integral usando la Regla del Trapecio compuesta.")
ws.cell(row=22, column=2, value="2. Las columnas 'k=2' a 'k=5' aplican la Extrapolación de Richardson para reducir el error de truncamiento de O(h^2) a O(h^10).")
ws.cell(row=23, column=2, value="3. El resultado final (la aproximación más exacta) se encuentra en la celda K2 resaltada en verde.")
ws.cell(row=24, column=2, value="Puedes alterar los valores de f(t) en la columna C y ver cómo toda la matriz de Romberg se recalcula automáticamente gracias a las fórmulas.")

ws.cell(row=2, column=11).fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
ws.cell(row=2, column=11).font = Font(bold=True, color="FFFFFF")

wb.save("Romberg_Contraste.xlsx")
