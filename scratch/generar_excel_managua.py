import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import numpy as np
import sys
import os

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from aplicacion.clima_api import obtener_datos_clima
from aplicacion.optimizacion import calcular_angulo_optimo

lat = 12.1462
lon = -86.2782
coef = -0.45

# Obtener datos del clima
clima = obtener_datos_clima(lat, lon)
if not clima.get('success'):
    print("Error:", clima)
    sys.exit(1)

datos = clima["datos_horarios"]
# Obtener radiación real con pérdida de temperatura
y_vals_real = []
for d in datos:
    rad = d["radiacion"]
    temp = d.get("temperatura", 25.0)
    if temp > 25.0:
        perdida = abs(coef) / 100.0 * (temp - 25.0)
        y_vals_real.append(rad * (1.0 - perdida))
    else:
        y_vals_real.append(rad)

# Para interpolación en Romberg (si h no es entero)
x_vals = np.arange(len(y_vals_real))
def f_real(t):
    return float(np.interp(t, x_vals, y_vals_real))

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Metodo de Romberg"

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
align_center = Alignment(horizontal="center", vertical="center")

# Añadimos los encabezados
headers = [
    "i", "t (Hora)", "f(t) (W/m²)", 
    "j (Nivel)", "Intervalos (n)", "h (Paso)", "k=1 (Trapecio)", 
    "k=2", "k=3", "k=4", "k=5 (Alta Precisión)"
]

ws.append(headers)

for col in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = align_center

# Romberg en el sistema usa n_trap = 2**(j+1) para j de 0 a 4.
# O sea, n = 2, 4, 8, 16, 32.
# El máximo n es 32, así que ocupamos 33 puntos de datos.
n = 32
a = 0
b = 23
h_min = (b - a) / n

# Generar los 33 puntos (i = 0 a 32)
for i in range(33):
    t_val = a + i * h_min
    ft = f_real(t_val)
    ws.append([i, t_val, ft])

# Formato de números para puntos
for row in range(2, 35):
    ws.cell(row=row, column=2).number_format = '0.0000'
    ws.cell(row=row, column=3).number_format = '0.0000'

# Tabla de Romberg (inicia en fila 2, columna 4)
# Niveles 1 a 5 (corresponden a j=0 a 4 en Python)
for j in range(1, 6):
    row = j + 1
    n_j = 2 ** j  # 2, 4, 8, 16, 32
    h_j = (b - a) / n_j
    ws.cell(row=row, column=4, value=j)
    ws.cell(row=row, column=5, value=n_j)
    ws.cell(row=row, column=6, value=h_j)
    
    # Fórmulas del trapecio referenciando la columna C (C2 a C34)
    # i=0 es C2, i=32 es C34.
    if j == 1: # n=2 (saltos de 16, i=0, 16, 32)
        f_trap = f"=(F{row}/2)*(C2 + 2*C18 + C34)"
    elif j == 2: # n=4 (saltos de 8, i=0, 8, 16, 24, 32)
        f_trap = f"=(F{row}/2)*(C2 + 2*(C10+C18+C26) + C34)"
    elif j == 3: # n=8 (saltos de 4)
        f_trap = f"=(F{row}/2)*(C2 + 2*(C6+C10+C14+C18+C22+C26+C30) + C34)"
    elif j == 4: # n=16 (saltos de 2)
        f_trap = f"=(F{row}/2)*(C2 + 2*(C4+C6+C8+C10+C12+C14+C16+C18+C20+C22+C24+C26+C28+C30+C32) + C34)"
    elif j == 5: # n=32 (todos)
        f_trap = f"=(F{row}/2)*(C2 + 2*SUM(C3:C33) + C34)"
        
    ws.cell(row=row, column=7, value=f_trap)

cols = ['G', 'H', 'I', 'J', 'K']

for k in range(2, 6):
    col_idx = 6 + k
    factor = 4 ** (k - 1)
    
    for j in range(1, 6 - k + 1):
        row = j + 1
        prev_col = cols[k - 2]
        f_extrap = f"=({factor}*{prev_col}{row + 1} - {prev_col}{row})/{factor - 1}"
        ws.cell(row=row, column=col_idx, value=f_extrap)

# Dar formato a tabla Romberg
for row in range(2, 7):
    ws.cell(row=row, column=6).number_format = '0.0000'
    for col in range(7, 12):
        ws.cell(row=row, column=col).number_format = '0.0000'

for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
    ws.column_dimensions[col].width = 16

ws.cell(row=2, column=11).fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
ws.cell(row=2, column=11).font = Font(bold=True, color="FFFFFF")


# === CÁLCULO TRAPECIO SIMPLE DEL SISTEMA (h=1, 23 intervalos) ===
# Se hace en M1
ws.cell(row=2, column=13, value="Trapecio Simple (h=1) como en Sistema").font = Font(bold=True)
ws.cell(row=3, column=13, value="El sistema evalúa el trapecio usando los 24 puntos de datos originales (h=1).")
# No podemos usar fórmulas directas a los datos de la columna C porque la columna C es para n=32 (h=0.71875)
# Calculamos el valor fijo aquí para referencia
h_trap = 1
suma_interna = sum(y_vals_real[1:-1])
integral_trap_simple = (h_trap / 2) * (y_vals_real[0] + 2*suma_interna + y_vals_real[-1])

ws.cell(row=4, column=13, value="Energía (Trapecio Simple):")
ws.cell(row=4, column=14, value=integral_trap_simple).number_format = '0.0000'
ws.cell(row=5, column=13, value="Energía (Romberg k=5):")
ws.cell(row=5, column=14, value="=K2").number_format = '0.0000'

# CREAR HOJA PARA NEWTON-RAPHSON
ws2 = wb.create_sheet("Newton-Raphson")

headers_nr = ["Iteración", "xi", "f(xi)", "f'(xi)", "xi+1", "Error"]
ws2.append(headers_nr)

for col in range(1, len(headers_nr) + 1):
    cell = ws2.cell(row=1, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = align_center

max_rad = clima["max_radiacion"]
optimizacion = calcular_angulo_optimo(lat, max_rad)

if optimizacion and "pasos" in optimizacion:
    for paso in optimizacion["pasos"]:
        ws2.append([
            paso["iter"],
            paso["ci"],
            paso["fci"],
            paso["dfci"],
            paso["cimas1"],
            paso["error"]
        ])

for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    ws2.column_dimensions[col].width = 18

for row in range(2, ws2.max_row + 1):
    for col in range(2, 7):
        ws2.cell(row=row, column=col).number_format = '0.00000'

output_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'Calculos_Managua_v2.xlsx'))
wb.save(output_path)
print(f"Archivo guardado en {output_path}")
