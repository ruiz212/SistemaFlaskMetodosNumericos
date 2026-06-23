import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import numpy as np
import sys
import os

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from aplicacion.clima_api import obtener_datos_clima

lat = 12.1462
lon = -86.2782
coef = -0.45

# Obtener datos del clima
clima = obtener_datos_clima(lat, lon)
if not clima.get('success'):
    print("Error:", clima)
    sys.exit(1)

datos = clima["datos_horarios"]
# Obtener radiación real con pérdida de temperatura
y_vals_real = []
for d in datos:
    rad = d["radiacion"]
    temp = d.get("temperatura", 25.0)
    if temp > 25.0:
        perdida = abs(coef) / 100.0 * (temp - 25.0)
        y_vals_real.append(rad * (1.0 - perdida))
    else:
        y_vals_real.append(rad)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Metodo del Trapecio Simple"

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
align_center = Alignment(horizontal="center", vertical="center")

# Añadimos los encabezados
headers = [
    "i", "t (Hora)", "Radiación Real f(t) (W/m²)", "Tipo de Punto", "Multiplicador en Fórmula", "Contribución a Suma"
]

ws.append(headers)

for col in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = align_center

n = 23 # 24 puntos - 1
a = 0
b = 23
h = 1 # hora

# Generar los 24 puntos
for i in range(24):
    t_val = a + i * h
    ft = y_vals_real[i]
    
    tipo = ""
    mult = 1
    if i == 0:
        tipo = "Extremo f(x0)"
    elif i == 23:
        tipo = "Extremo f(xn)"
    else:
        tipo = "Punto interno f(xi)"
        mult = 2
        
    ws.append([i, t_val, ft, tipo, mult, f"=C{i+2}*E{i+2}"])

# Formato numérico
for row in range(2, 26):
    ws.cell(row=row, column=2).number_format = '0'
    ws.cell(row=row, column=3).number_format = '0.00'
    ws.cell(row=row, column=6).number_format = '0.00'

# Sumatoria total
ws.cell(row=27, column=5, value="Suma Total (Corchete):").font = Font(bold=True)
ws.cell(row=27, column=6, value="=SUM(F2:F25)").number_format = '0.00'

ws.cell(row=28, column=5, value="h (Paso en horas):").font = Font(bold=True)
ws.cell(row=28, column=6, value=h)

ws.cell(row=29, column=5, value="Energía Total (Wh):").font = Font(bold=True)
ws.cell(row=29, column=6, value="=(F28/2)*F27").number_format = '0.00'
ws.cell(row=29, column=6).fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
ws.cell(row=29, column=6).font = Font(bold=True, color="FFFFFF")

# Ajustar ancho columnas
for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    ws.column_dimensions[col].width = 25

output_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'Trapecio_Managua.xlsx'))
wb.save(output_path)
print(f"Archivo guardado en {output_path}")
