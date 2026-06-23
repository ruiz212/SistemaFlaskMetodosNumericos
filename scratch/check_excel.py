import openpyxl

wb = openpyxl.load_workbook("Calculos_Managua.xlsx", data_only=False)
ws = wb["Metodo de Romberg"]

# We can't evaluate formulas easily in pure python without xlwings or win32com, 
# but let's check the sum of trapecio manually or just see if the input data matches the system.
import numpy as np

# Print the formula in K2
print("Formula K2:", ws["K2"].value)

# Let's calculate the Romberg result in Python using the exact same data as the excel sheet
y_vals = []
for row in range(2, 19): # 17 points
    y_vals.append(float(ws.cell(row=row, column=3).value))

print("y_vals in Excel:", y_vals)
print("Sum y_vals[1:-1]:", sum(y_vals[1:-1]))
print("y_vals[0]:", y_vals[0])
print("y_vals[-1]:", y_vals[-1])

h = 23.0 / 16.0
trap = (h / 2) * (y_vals[0] + 2 * sum(y_vals[1:-1]) + y_vals[-1])
print("Trapecio k=1 (Excel data):", trap)

# But wait, in the system, Romberg uses len(x_vals)-1 = 23 intervals!
# len(x_vals) = 24.
# `res_romberg = romberg(0, len(x_vals)-1, 5, f_real)`
# In `romberg` function:
# `n_trap = 2**(j+1)`
# So for j=0, n_trap=2. j=1, n_trap=4. j=2, n_trap=8. j=3, n_trap=16. j=4, n_trap=32.
# So Romberg uses 32 intervals!
# But my Excel generator `generar_excel_managua.py` uses `n = 16` intervals for the grid.
# Wait, `generar_excel_managua.py` generates the Excel grid with 17 points, and the formula uses those 17 points.
# BUT `romberg(0, 23, 5, f_real)` uses `2**j * 2` intervals if `j` goes up to 5!
# Actually, in `metodos/integracion.py`, `n_trap = 2**(j+1)`. So for 5 niveles (j=0,1,2,3,4):
# n_trap = 2, 4, 8, 16, 32.
# In `generar_excel_managua.py`, I hardcoded `n = 16` and formulas that only go up to 16 intervals!
# Ah! In `generar_excel_managua.py`:
# `for j in range(1, 6): n_j = 2 ** (j - 1)`
# So n_j = 1, 2, 4, 8, 16.
# That is DIFFERENT from `n_trap = 2**(j+1)` (2, 4, 8, 16, 32) in `metodos/integracion.py`!
# Let's verify `metodos/integracion.py` lines 56-57:
# `n_trap = 2**(j+1)`
